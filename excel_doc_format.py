from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def gerar_planilha_excel_com_formatacao():
    # Criando um novo arquivo Excel
    wb = Workbook()
    planilha = wb.active

    # Adicionando dados à planilha
    planilha["A1"] = "Exemplo de Planilha Excel com Formatação"
    planilha.merge_cells("A1:D1")
    planilha["A2"] = "Item"
    planilha["B2"] = "Quantidade"
    planilha["C2"] = "Preço Unitário"
    planilha["D2"] = "Total"
    dados = [("Produto 1", 5, 10.0), ("Produto 2", 10, 15.0), ("Produto 3", 3, 20.0)]
    for linha, (produto, quantidade, preco) in enumerate(dados, start=3):
        planilha[f"A{linha}"] = produto
        planilha[f"B{linha}"] = quantidade
        planilha[f"C{linha}"] = preco
        planilha[f"D{linha}"] = quantidade * preco

    # Aplicando formatações
    titulo_fonte = Font(size=14, bold=True)
    planilha["A1"].font = titulo_fonte
    planilha["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for col in planilha.iter_cols(min_row=2, max_row=2, min_col=1, max_col=4):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    for row in planilha.iter_rows(min_row=3, min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Salvando o arquivo
    wb.save("planilha_excel_com_formatacao.xlsx")

if __name__ == "__main__":
    gerar_planilha_excel_com_formatacao()